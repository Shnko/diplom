import tempfile
from pathlib import Path
from uuid import uuid4

import openpyxl
import random

from django.db.models import Q, Sum

from core.models import Orphanage, Employment, ChildAdmission, ChildReturned, ChildCare, ChildAdopted, \
    ChildRepatriation, InternationalAdoption, TransferToTreatment, TransferByCertainAge, ChildDeath, PositionCategory, \
    PlannedRates
from orphanage.settings import BASE_DIR


def build_report_2120(start, end):
    # исходный файл с незаполненной таблицей
    source = BASE_DIR / 'core' / 'templates' / 'reports' / 'report_2120.xlsx'

    # файл назначения
    destination = Path(tempfile.mkdtemp()) / str(uuid4().hex)

    # вывод в консоль
    print(f'Report 2120 to "{destination}" from "{start}" to "{end}"')

    # открытие файла
    workbook = openpyxl.load_workbook(source)

    # получение листа
    spreadsheet = workbook.active

    # заполнение данных в таблицу
    for i in range(5, 8):
        for j in range(3, 12):
            cell = spreadsheet.cell(row=i, column=j)
            cell.value = random.randint(-2000, 2000) / 1000

    # сохранение файла во временный файл
    workbook.save(destination)
    workbook.close()
    return destination


def build_report_summary(start, end):
    # исходный файл с незаполненной таблицей
    source = BASE_DIR / 'core' / 'templates' / 'reports' / 'report_summary.xlsx'

    # файл назначения
    destination = Path(tempfile.mkdtemp()) / str(uuid4().hex)

    # вывод в консоль
    print(f'Report 2120 to "{destination}" from "{start}" to "{end}"')

    # открытие файла
    workbook = openpyxl.load_workbook(source)

    # получение листа
    spreadsheet = workbook.active

    # заполнение таблицы данными 1 Дом ребёнка
    spreadsheet["A6"].value = 1
    orphanage = Orphanage.objects.first()
    spreadsheet["B6"].value = 1
    spreadsheet["F6"].value = orphanage.count_of_seats

    # заполнение таблицы данными 2 Штаты организации, человек

    plan_teachers = PlannedRates.objects.filter(category='6').aggregate(Sum('count'))['count__sum']
    plan_others = PlannedRates.objects.filter(category='8').aggregate(Sum('count'))['count__sum']

    spreadsheet["C16"].value = PlannedRates.objects.aggregate(Sum('count'))['count__sum']
    spreadsheet["D16"].value = PlannedRates.objects.filter(category='1').aggregate(Sum('count'))['count__sum']
    spreadsheet["G16"].value = PlannedRates.objects.filter(category='2').aggregate(Sum('count'))['count__sum']
    spreadsheet["J16"].value = plan_others + plan_teachers
    spreadsheet["M16"].value = plan_teachers

    spreadsheet["C17"].value = Employment.objects.filter(
        Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end))).count()

    spreadsheet["D17"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.DOCTOR) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["E17"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.NON_MEDICAL_STAFF) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["F17"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.PHARMACIST) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["G17"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.LINEAR_MEDICAL_STAFF) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["H17"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.APOTHECARY) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["I17"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.JUNIOR_MEDICAL_STAFF) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    teachers = Employment.objects.filter(
        Q(report_category=PositionCategory.TEACHER) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    others = Employment.objects.filter(
        Q(report_category=PositionCategory.OTHERS) &
        (Q(end_date_of_employment__isnull=True) | Q(end_date_of_employment__range=(start, end)))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["J17"].value = others + teachers

    spreadsheet["M17"].value = teachers

    spreadsheet["C18"].value = Employment.objects.filter(
        Q(end_date_of_employment__range=(start, end))).count()

    spreadsheet["D18"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.DOCTOR) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["E18"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.NON_MEDICAL_STAFF) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["F18"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.PHARMACIST) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["G18"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.LINEAR_MEDICAL_STAFF) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["H18"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.APOTHECARY) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["I18"].value = Employment.objects.filter(
        Q(report_category=PositionCategory.JUNIOR_MEDICAL_STAFF) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    teachers = Employment.objects.filter(
        Q(report_category=PositionCategory.TEACHER) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    others = Employment.objects.filter(
        Q(report_category=PositionCategory.OTHERS) &
        Q(end_date_of_employment__range=(start, end))).aggregate(
        Sum('rate'))['rate__sum']

    spreadsheet["J18"].value = others + teachers

    spreadsheet["M18"].value = teachers

    # заполнение таблицы данными 3 Контингенты дома ребенка, человек
    orphan_child = ChildAdmission.objects.filter(
        Q(admission_type=ChildAdmission.AdmissionType.ORPHAN
          ) & Q(date_of_admission__range=(start, end))).count()

    without_care_child = ChildAdmission.objects.filter(
        Q(admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_admission__range=(start, end))).count()

    spreadsheet["C29"].value = ChildAdmission.objects.filter(Q(date_of_admission__range=(start, end))).count()
    spreadsheet["C30"].value = without_care_child
    spreadsheet["C31"].value = orphan_child

    orphan_admission_type_query =  Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.ORPHAN
          ) & Q(date_of_adoption__range=(start, end))

    # ORPHAN
    orphan_child_care_count = ChildCare.objects.filter(orphan_admission_type_query).count()

    orphan_child_adopted_count = ChildAdopted.objects.filter(orphan_admission_type_query).count()

    orphan_child_returned_count = ChildReturned.objects.filter(orphan_admission_type_query).count()

    orphan_child_international_adoption_count = InternationalAdoption.objects.filter(orphan_admission_type_query).count()

    orphan_child_repatriation_count = ChildRepatriation.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.ORPHAN
          ) & Q(date_of_repatriation__range=(start, end))).count()

    orphan_transfer_to_treatment_count = TransferToTreatment.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.ORPHAN
          ) & Q(date_of_transfer__range=(start, end))).count()

    orphan_transfer_by_age_count = TransferByCertainAge.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.ORPHAN
          ) & Q(date_of_transfer__range=(start, end))).count()

    spreadsheet["D31"].value = orphan_child_care_count + orphan_child_adopted_count + orphan_child_returned_count + orphan_child_international_adoption_count + orphan_child_repatriation_count + orphan_transfer_to_treatment_count + orphan_transfer_by_age_count

    # WITHOUT_CARE
    care_child_care_count = ChildCare.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_adoption__range=(start, end))).count()

    care_child_adopted_count = ChildAdopted.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_adoption__range=(start, end))).count()

    care_child_returned_count = ChildReturned.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_adoption__range=(start, end))).count()

    care_child_international_adoption_count = InternationalAdoption.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_adoption__range=(start, end))).count()

    care_child_repatriation_count = ChildRepatriation.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_repatriation__range=(start, end))).count()

    care_transfer_to_treatment_count = TransferToTreatment.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_transfer__range=(start, end))).count()

    care_transfer_by_age_count = TransferByCertainAge.objects.filter(
        Q(child__childadmission__admission_type=ChildAdmission.AdmissionType.WITHOUT_CARE
          ) & Q(date_of_transfer__range=(start, end))).count()

    spreadsheet["D30"].value = (care_child_care_count + care_child_adopted_count + care_child_returned_count +
                                care_child_international_adoption_count + care_child_repatriation_count +
                                care_child_repatriation_count + care_transfer_to_treatment_count +
                                care_transfer_by_age_count)

    spreadsheet["F29"].value = ChildDeath.objects.count()

    # заполнение таблицы данными 4 Движение контингентов дома ребенка, человек

    spreadsheet["A42"].value = ChildReturned.objects.filter(
        Q(date_of_adoption__range=(start, end))).count()

    spreadsheet["B42"].value = ChildAdopted.objects.filter(
        Q(date_of_adoption__range=(start, end))).count()

    spreadsheet["D42"].value = TransferByCertainAge.objects.filter(
        Q(type=TransferByCertainAge.InstitutionType.EDUCATION_INST) &
        Q(date_of_transfer__range=(start, end))).count()

    spreadsheet["E42"].value = TransferByCertainAge.objects.filter(
        Q(type=TransferByCertainAge.InstitutionType.SOCIAL_SAFE_INST) &
        Q(date_of_transfer__range=(start, end))).count()

    spreadsheet["F42"].value = InternationalAdoption.objects.filter(
        Q(date_of_adoption__range=(start, end))).count()

    spreadsheet["H42"].value = ChildCare.objects.filter(
        Q(care_type=ChildCare.CareType.WARDERED) &
        Q(date_of_adoption__range=(start, end))).count()

    spreadsheet["J42"].value = ChildCare.objects.filter(
        Q(care_type=ChildCare.CareType.FOSTER_FAMILY) &
        Q(date_of_adoption__range=(start, end))).count()

    spreadsheet["L42"].value = ChildRepatriation.objects.filter(
        Q(date_of_repatriation__range=(start, end))).count()

    spreadsheet["M42"].value = TransferToTreatment.objects.filter(
        Q(date_of_transfer__range=(start, end))).count()

    # заполнение таблицы данными 5 Профилактические осмотри

    # сохранение файла во временный файл
    workbook.save(destination)
    workbook.close()
    return destination


if __name__ == '__main__':
    build_report_2120(1, 2)
